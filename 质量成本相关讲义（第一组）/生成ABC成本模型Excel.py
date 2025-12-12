# -*- coding: utf-8 -*-
"""
瓦轴集团ABC成本模型 - Excel生成器
基于完整模拟数据集生成可直接使用的Excel模型
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# 创建工作簿
wb = openpyxl.Workbook()

# 定义样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
calc_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
title_font = Font(name="微软雅黑", size=14, bold=True)
normal_font = Font(name="微软雅黑", size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_width(ws, col, width):
    """设置列宽"""
    ws.column_dimensions[get_column_letter(col)].width = width

def format_header(ws, row, start_col, end_col):
    """格式化表头行"""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

# ============================================================
# 工作表1: 说明
# ============================================================
ws1 = wb.active
ws1.title = "说明"

# 标题
ws1['B2'] = "瓦轴集团ABC成本核算模型"
ws1['B2'].font = Font(name="微软雅黑", size=18, bold=True, color="4472C4")
ws1['B3'] = "Activity-Based Costing Model"
ws1['B3'].font = Font(name="微软雅黑", size=12, italic=True, color="7F7F7F")

# 基本信息
ws1['B5'] = "模型版本："
ws1['C5'] = "V1.0"
ws1['B6'] = "创建日期："
ws1['C6'] = datetime.date.today().strftime("%Y年%m月%d日")
ws1['B7'] = "适用范围："
ws1['C7'] = "精加工车间三分厂"
ws1['B8'] = "核算期间："
ws1['C8'] = "2024年第四季度(10-12月)"

# 使用说明
ws1['B10'] = "使用说明："
ws1['B10'].font = Font(name="微软雅黑", size=12, bold=True)
instructions = [
    "1. 首先查看【基础数据】工作表，了解产品和生产信息",
    "2. 查看【成本归集】工作表，了解制造费用构成",
    "3. 查看【作业识别】和【成本动因】，了解ABC方法的核心",
    "4. 系统自动计算【产品成本】，无需手动操作",
    "5. 查看【成本对比】和【可视化图表】，了解ABC方法的价值"
]
for i, text in enumerate(instructions):
    ws1[f'B{11+i}'] = text
    ws1[f'B{11+i}'].font = normal_font

# 注意事项
ws1['B17'] = "注意事项："
ws1['B17'].font = Font(name="微软雅黑", size=12, bold=True)
notes = [
    "• 浅黄色单元格为输入区（本模型为演示，数据已填充）",
    "• 浅蓝色单元格为自动计算区，请勿修改",
    "• 修改输入数据后，按F9刷新计算",
    "• 定期备份模型文件"
]
for i, text in enumerate(notes):
    ws1[f'B{18+i}'] = text
    ws1[f'B{18+i}'].font = normal_font

# 项目组信息
ws1['B23'] = "项目组成员："
ws1['B23'].font = Font(name="微软雅黑", size=11, bold=True)
ws1['B24'] = "负责人: __________"
ws1['B25'] = "成员: __________, __________, __________"

ws1['B27'] = "技术支持："
ws1['B27'].font = Font(name="微软雅黑", size=11, bold=True)
ws1['B28'] = "顾问: __________  联系方式: __________"

# 设置列宽
set_column_width(ws1, 1, 3)
set_column_width(ws1, 2, 25)
set_column_width(ws1, 3, 40)

# ============================================================
# 工作表2: 基础数据
# ============================================================
ws2 = wb.create_sheet("基础数据")

# 产品信息表
ws2['A1'] = "产品信息表"
ws2['A1'].font = title_font
ws2.merge_cells('A1:I1')

headers = ["产品编号", "产品型号", "产品名称", "产品类别", "季度产量(件)",
           "批次数", "平均批量", "单位售价(元)", "备注"]
for i, header in enumerate(headers, 1):
    cell = ws2.cell(row=2, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 产品数据
products = [
    ["P001", "32315", "圆锥滚子轴承", "标准品", 60000, 30, 2000, 145, "主力产品,走量"],
    ["P002", "6320", "深沟球轴承", "标准品", 45000, 45, 1000, 92, "薄利多销"],
    ["P003", "NU2316", "圆柱滚子轴承", "标准品", 18000, 36, 500, 280, "毛利较好"],
    ["P004", "定制-HD", "高端定制轴承", "定制品", 1500, 30, 50, 1800, "技术门槛高"],
    ["P005", "定制-TP", "特种工况轴承", "定制品", 300, 20, 15, 5500, "超高端定制"],
]

for row_idx, product in enumerate(products, 3):
    for col_idx, value in enumerate(product, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = normal_font
        if col_idx >= 5 and col_idx <= 8:  # 数值列
            cell.alignment = Alignment(horizontal='right')
            if col_idx == 5 or col_idx == 6 or col_idx == 7:
                cell.number_format = '#,##0'
            elif col_idx == 8:
                cell.number_format = '#,##0.00'

# 合计行
total_row = 8
ws2[f'A{total_row}'] = "合计"
ws2[f'A{total_row}'].font = Font(name="微软雅黑", size=10, bold=True)
ws2[f'E{total_row}'] = f"=SUM(E3:E7)"
ws2[f'F{total_row}'] = f"=SUM(F3:F7)"
ws2[f'E{total_row}'].number_format = '#,##0'
ws2[f'F{total_row}'].number_format = '#,##0'
for col in range(1, 10):
    ws2.cell(row=total_row, column=col).border = thin_border
    ws2.cell(row=total_row, column=col).fill = calc_fill

# 工时统计表
ws2['A10'] = "产品工时统计表"
ws2['A10'].font = title_font
ws2.merge_cells('A10:E10')

headers2 = ["产品编号", "单件标准工时(h)", "单件机器小时(h)", "季度总人工(h)", "季度总机时(h)"]
for i, header in enumerate(headers2, 1):
    cell = ws2.cell(row=11, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

workhours = [
    ["P001", 0.4, 1.2, 24000, 72000],
    ["P002", 0.3, 0.8, 13500, 36000],
    ["P003", 0.6, 1.5, 10800, 27000],
    ["P004", 2.5, 4.0, 3750, 6000],
    ["P005", 5.0, 8.0, 1500, 2400],
]

for row_idx, wh in enumerate(workhours, 12):
    for col_idx, value in enumerate(wh, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = normal_font
        if col_idx >= 2:
            cell.alignment = Alignment(horizontal='right')
            if col_idx >= 4:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '0.0'

# 合计行
total_row2 = 17
ws2[f'A{total_row2}'] = "合计"
ws2[f'A{total_row2}'].font = Font(name="微软雅黑", size=10, bold=True)
ws2[f'D{total_row2}'] = f"=SUM(D12:D16)"
ws2[f'E{total_row2}'] = f"=SUM(E12:E16)"
ws2[f'D{total_row2}'].number_format = '#,##0'
ws2[f'E{total_row2}'].number_format = '#,##0'
for col in range(1, 6):
    ws2.cell(row=total_row2, column=col).border = thin_border
    ws2.cell(row=total_row2, column=col).fill = calc_fill

# 直接成本汇总表
ws2['A19'] = "直接成本汇总表"
ws2['A19'].font = title_font
ws2.merge_cells('A19:E19')

headers3 = ["产品编号", "直接材料(元)", "直接人工(元)", "直接成本合计(元)", "单位直接成本(元)"]
for i, header in enumerate(headers3, 1):
    cell = ws2.cell(row=20, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

direct_costs = [
    ["P001", 2880000, 1200000],
    ["P002", 1170000, 675000],
    ["P003", 1170000, 540000],
    ["P004", 210000, 300000],
    ["P005", 90000, 120000],
]

for row_idx, dc in enumerate(direct_costs, 21):
    ws2.cell(row=row_idx, column=1, value=dc[0]).border = thin_border
    ws2.cell(row=row_idx, column=2, value=dc[1]).border = thin_border
    ws2.cell(row=row_idx, column=3, value=dc[2]).border = thin_border
    # 直接成本合计
    ws2.cell(row=row_idx, column=4, value=f"=B{row_idx}+C{row_idx}").border = thin_border
    # 单位直接成本
    prod_row = row_idx - 18
    ws2.cell(row=row_idx, column=5, value=f"=D{row_idx}/E{prod_row}").border = thin_border

    for col in range(2, 6):
        ws2.cell(row=row_idx, column=col).font = normal_font
        ws2.cell(row=row_idx, column=col).number_format = '#,##0.00'
        ws2.cell(row=row_idx, column=col).alignment = Alignment(horizontal='right')

# 设置列宽
for col in range(1, 10):
    set_column_width(ws2, col, 15)

# ============================================================
# 工作表3: 成本归集
# ============================================================
ws3 = wb.create_sheet("成本归集")

ws3['A1'] = "制造费用汇总表（季度，元）"
ws3['A1'].font = title_font
ws3.merge_cells('A1:F1')

headers4 = ["费用编号", "费用科目", "季度发生额(元)", "占比", "归属性质", "备注"]
for i, header in enumerate(headers4, 1):
    cell = ws3.cell(row=2, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

overhead_data = [
    ["C01", "设备折旧费", 2400000, "与设备使用相关"],
    ["C02", "电费", 900000, "与机器运行相关"],
    ["C03", "设备维修保养费", 480000, "与设备使用相关"],
    ["C04", "工装模具折旧", 360000, "与批次相关"],
    ["C05", "车间管理人员工资", 1260000, "设施级"],
    ["C06", "间接生产人员工资", 840000, "多种作业"],
    ["C07", "质检部门费用", 720000, "与检验相关"],
    ["C08", "物料搬运费用", 360000, "与搬运相关"],
    ["C09", "车间办公及低耗", 180000, "设施级"],
    ["C10", "水费、蒸汽费", 240000, "与生产相关"],
    ["C11", "其他制造费用", 160000, "设施级"],
]

for row_idx, od in enumerate(overhead_data, 3):
    ws3.cell(row=row_idx, column=1, value=od[0]).border = thin_border
    ws3.cell(row=row_idx, column=2, value=od[1]).border = thin_border
    ws3.cell(row=row_idx, column=3, value=od[2]).border = thin_border
    # 占比公式
    ws3.cell(row=row_idx, column=4, value=f"=C{row_idx}/C14").border = thin_border
    ws3.cell(row=row_idx, column=4).number_format = '0.0%'
    ws3.cell(row=row_idx, column=5, value=od[3]).border = thin_border

    for col in range(1, 7):
        ws3.cell(row=row_idx, column=col).font = normal_font
    ws3.cell(row=row_idx, column=3).number_format = '#,##0'
    ws3.cell(row=row_idx, column=3).alignment = Alignment(horizontal='right')

# 合计行
total_row3 = 14
ws3[f'A{total_row3}'] = "合计"
ws3[f'A{total_row3}'].font = Font(name="微软雅黑", size=10, bold=True)
ws3[f'C{total_row3}'] = f"=SUM(C3:C13)"
ws3[f'C{total_row3}'].number_format = '#,##0'
ws3[f'D{total_row3}'] = "100.0%"
for col in range(1, 7):
    ws3.cell(row=total_row3, column=col).border = thin_border
    ws3.cell(row=total_row3, column=col).fill = calc_fill

# 设置列宽
set_column_width(ws3, 1, 10)
set_column_width(ws3, 2, 20)
set_column_width(ws3, 3, 15)
set_column_width(ws3, 4, 10)
set_column_width(ws3, 5, 20)
set_column_width(ws3, 6, 15)

# ============================================================
# 工作表4: 作业识别
# ============================================================
ws4 = wb.create_sheet("作业识别")

ws4['A1'] = "作业清单"
ws4['A1'].font = title_font
ws4.merge_cells('A1:G1')

headers5 = ["作业编号", "作业名称", "作业层级", "作业描述", "作业成本(元)", "占比", "备注"]
for i, header in enumerate(headers5, 1):
    cell = ws4.cell(row=2, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

activities = [
    ["A01", "车削加工", "单位级", "内外圈粗精车", 1680000],
    ["A02", "磨削加工", "单位级", "内外圈精密磨", 1440000],
    ["A03", "热处理", "单位级", "淬火回火", 720000],
    ["A04", "超精研", "单位级", "表面超精加工", 480000],
    ["A05", "清洗去毛刺", "单位级", "清洗处理", 240000],
    ["B01", "设备换型调整", "批次级", "工装更换调整", 800000],
    ["B02", "首件检验", "批次级", "批次首检", 240000],
    ["B03", "生产准备", "批次级", "领料排产", 160000],
    ["B04", "物料搬运", "批次级", "工序间搬运", 360000],
    ["B05", "批次质检", "批次级", "巡检抽检", 320000],
    ["B06", "包装入库", "批次级", "批次包装", 160000],
    ["C01", "工艺设计优化", "产品级", "新品工艺", 240000],
    ["C02", "专用工装制作", "产品级", "专用工装", 180000],
    ["C03", "程序编制调试", "产品级", "数控程序", 120000],
    ["C04", "试产验证", "产品级", "新品试产", 80000],
    ["D01", "车间管理", "设施级", "车间运营管理", 480000],
    ["D02", "设备日常维护", "设施级", "预防性维护", 200000],
    ["D03", "质量体系维护", "设施级", "质量管理", 160000],
    ["D04", "环境安全管理", "设施级", "5S安全", 120000],
    ["D05", "能源动力供应", "设施级", "水电气供应", 120000],
]

for row_idx, act in enumerate(activities, 3):
    for col_idx, value in enumerate(act, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.font = normal_font
        if col_idx == 5:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')

# 占比公式
for row_idx in range(3, 23):
    ws4.cell(row=row_idx, column=6, value=f"=E{row_idx}/E23")
    ws4.cell(row=row_idx, column=6).number_format = '0.0%'
    ws4.cell(row=row_idx, column=6).border = thin_border

# 合计行
ws4['A23'] = "合计"
ws4['A23'].font = Font(name="微软雅黑", size=10, bold=True)
ws4['E23'] = "=SUM(E3:E22)"
ws4['E23'].number_format = '#,##0'
ws4['F23'] = "100.0%"
for col in range(1, 8):
    ws4.cell(row=23, column=col).border = thin_border
    ws4.cell(row=23, column=col).fill = calc_fill

# 设置列宽
set_column_width(ws4, 1, 10)
set_column_width(ws4, 2, 18)
set_column_width(ws4, 3, 10)
set_column_width(ws4, 4, 18)
set_column_width(ws4, 5, 15)
set_column_width(ws4, 6, 10)
set_column_width(ws4, 7, 15)

print("工作表 '作业识别' 创建完成...")

# ============================================================
# 工作表5: 成本动因
# ============================================================
ws5 = wb.create_sheet("成本动因")

ws5['A1'] = "成本动因选择与分配率"
ws5['A1'].font = title_font
ws5.merge_cells('A1:G1')

headers6 = ["作业编号", "作业名称", "成本动因", "动因总量", "作业成本(元)", "分配率", "单位"]
for i, header in enumerate(headers6, 1):
    cell = ws5.cell(row=2, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

cost_drivers = [
    ["A01", "车削加工", "车削机时(h)", 69000, 1680000, "元/h"],
    ["A02", "磨削加工", "磨削机时(h)", 61200, 1440000, "元/h"],
    ["A03", "热处理", "热处理件数", 124800, 720000, "元/件"],
    ["A04", "超精研", "超精研机时(h)", 13200, 480000, "元/h"],
    ["A05", "清洗去毛刺", "清洗件数", 124800, 240000, "元/件"],
    ["B01", "设备换型调整", "换型次数", 161, 800000, "元/次"],
    ["B02", "首件检验", "首检次数", 161, 240000, "元/次"],
    ["B03", "生产准备", "生产批次", 161, 160000, "元/批"],
    ["B04", "物料搬运", "搬运批次", 322, 360000, "元/批"],
    ["B05", "批次质检", "巡检批次", 161, 320000, "元/批"],
    ["B06", "包装入库", "包装批次", 161, 160000, "元/批"],
    ["C01", "工艺设计优化", "产品种类", 2, 240000, "元/种"],
    ["C02", "专用工装制作", "专用工装数", 5, 180000, "元/套"],
    ["C03", "程序编制调试", "程序套数", 5, 120000, "元/套"],
    ["C04", "试产验证", "试产次数", 4, 80000, "元/次"],
    ["D01", "车间管理", "产量(件)", 124800, 480000, "元/件"],
    ["D02", "设备日常维护", "机器小时", 143400, 200000, "元/h"],
    ["D03", "质量体系维护", "产量(件)", 124800, 160000, "元/件"],
    ["D04", "环境安全管理", "产量(件)", 124800, 120000, "元/件"],
    ["D05", "能源动力供应", "机器小时", 143400, 120000, "元/h"],
]

for row_idx, cd in enumerate(cost_drivers, 3):
    ws5.cell(row=row_idx, column=1, value=cd[0]).border = thin_border
    ws5.cell(row=row_idx, column=2, value=cd[1]).border = thin_border
    ws5.cell(row=row_idx, column=3, value=cd[2]).border = thin_border
    ws5.cell(row=row_idx, column=4, value=cd[3]).border = thin_border
    ws5.cell(row=row_idx, column=5, value=cd[4]).border = thin_border
    # 分配率公式
    ws5.cell(row=row_idx, column=6, value=f"=E{row_idx}/D{row_idx}")
    ws5.cell(row=row_idx, column=6).number_format = '#,##0.00'
    ws5.cell(row=row_idx, column=7, value=cd[5]).border = thin_border

    for col in range(1, 8):
        ws5.cell(row=row_idx, column=col).font = normal_font
        ws5.cell(row=row_idx, column=col).border = thin_border
    ws5.cell(row=row_idx, column=4).number_format = '#,##0'
    ws5.cell(row=row_idx, column=5).number_format = '#,##0'
    ws5.cell(row=row_idx, column=4).alignment = Alignment(horizontal='right')
    ws5.cell(row=row_idx, column=5).alignment = Alignment(horizontal='right')
    ws5.cell(row=row_idx, column=6).alignment = Alignment(horizontal='right')

# 设置列宽
set_column_width(ws5, 1, 10)
set_column_width(ws5, 2, 18)
set_column_width(ws5, 3, 18)
set_column_width(ws5, 4, 12)
set_column_width(ws5, 5, 15)
set_column_width(ws5, 6, 15)
set_column_width(ws5, 7, 10)

# ============================================================
# 工作表6: 产品成本(ABC)
# ============================================================
ws6 = wb.create_sheet("产品成本(ABC)")

ws6['A1'] = "产品完全成本汇总表（ABC方法）"
ws6['A1'].font = title_font
ws6.merge_cells('A1:K1')

headers7 = ["产品编号", "产品型号", "产量(件)", "直接材料", "直接人工",
            "ABC制造费用", "完全成本", "单位成本", "单位售价", "单位毛利", "毛利率"]
for i, header in enumerate(headers7, 1):
    cell = ws6.cell(row=2, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# ABC制造费用（从模拟数据集）
abc_overhead = [3168180, 2468205, 1382163, 596652, 284800]

product_costs = []
for idx, (prod, dc) in enumerate(zip(products, direct_costs)):
    row = [
        prod[0],  # 产品编号
        prod[1],  # 产品型号
        prod[4],  # 产量
        dc[1],    # 直接材料
        dc[2],    # 直接人工
        abc_overhead[idx],  # ABC制造费用
    ]
    product_costs.append(row)

for row_idx, pc in enumerate(product_costs, 3):
    for col_idx, value in enumerate(pc, 1):
        ws6.cell(row=row_idx, column=col_idx, value=value).border = thin_border
        ws6.cell(row=row_idx, column=col_idx).font = normal_font
        if col_idx >= 3:
            ws6.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'
            ws6.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='right')

# 添加计算列
for row_idx in range(3, 8):
    # 完全成本 = 直接材料 + 直接人工 + ABC制造费用
    ws6.cell(row=row_idx, column=7, value=f"=D{row_idx}+E{row_idx}+F{row_idx}")
    ws6.cell(row=row_idx, column=7).number_format = '#,##0.00'
    ws6.cell(row=row_idx, column=7).border = thin_border
    ws6.cell(row=row_idx, column=7).fill = calc_fill

    # 单位成本 = 完全成本 / 产量
    ws6.cell(row=row_idx, column=8, value=f"=G{row_idx}/C{row_idx}")
    ws6.cell(row=row_idx, column=8).number_format = '0.00'
    ws6.cell(row=row_idx, column=8).border = thin_border
    ws6.cell(row=row_idx, column=8).fill = calc_fill

    # 单位售价（从基础数据）
    prod_row = row_idx
    ws6.cell(row=row_idx, column=9, value=f"=基础数据!H{prod_row}")
    ws6.cell(row=row_idx, column=9).number_format = '0.00'
    ws6.cell(row=row_idx, column=9).border = thin_border

    # 单位毛利 = 单位售价 - 单位成本
    ws6.cell(row=row_idx, column=10, value=f"=I{row_idx}-H{row_idx}")
    ws6.cell(row=row_idx, column=10).number_format = '0.00'
    ws6.cell(row=row_idx, column=10).border = thin_border
    ws6.cell(row=row_idx, column=10).fill = calc_fill

    # 毛利率 = 单位毛利 / 单位售价
    ws6.cell(row=row_idx, column=11, value=f"=J{row_idx}/I{row_idx}")
    ws6.cell(row=row_idx, column=11).number_format = '0.0%'
    ws6.cell(row=row_idx, column=11).border = thin_border
    ws6.cell(row=row_idx, column=11).fill = calc_fill

# 合计行
ws6['A8'] = "合计"
ws6['A8'].font = Font(name="微软雅黑", size=10, bold=True)
for col, col_letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (7, 'G')]:
    ws6[f'{col_letter}8'] = f"=SUM({col_letter}3:{col_letter}7)"
    ws6[f'{col_letter}8'].number_format = '#,##0.00'
    ws6[f'{col_letter}8'].fill = calc_fill
    ws6[f'{col_letter}8'].border = thin_border

# 设置列宽
widths = [10, 12, 12, 12, 12, 15, 15, 12, 12, 12, 10]
for i, width in enumerate(widths, 1):
    set_column_width(ws6, i, width)

print("工作表 '产品成本(ABC)' 创建完成...")

# ============================================================
# 工作表7: 成本对比
# ============================================================
ws7 = wb.create_sheet("成本对比")

ws7['A1'] = "传统方法 vs ABC方法成本对比分析"
ws7['A1'].font = title_font
ws7.merge_cells('A1:G1')

# 单位成本对比
ws7['A3'] = "单位成本对比"
ws7['A3'].font = Font(name="微软雅黑", size=12, bold=True)

headers8 = ["产品型号", "传统方法(元)", "ABC方法(元)", "差异(元)", "差异率", "分析"]
for i, header in enumerate(headers8, 1):
    cell = ws7.cell(row=4, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 传统方法成本（按产量分摊）
traditional_costs = [131.30, 104.30, 158.30, 403.30, 763.17]
abc_costs = [120.80, 95.85, 171.79, 737.77, 1649.33]
analysis = ["被高估", "被高估", "被低估", "严重低估!", "极度低估!"]

for row_idx, (prod, trad, abc, ana) in enumerate(zip(products, traditional_costs, abc_costs, analysis), 5):
    ws7.cell(row=row_idx, column=1, value=prod[1]).border = thin_border
    ws7.cell(row=row_idx, column=2, value=trad).border = thin_border
    ws7.cell(row=row_idx, column=2).number_format = '0.00'
    ws7.cell(row=row_idx, column=3, value=abc).border = thin_border
    ws7.cell(row=row_idx, column=3).number_format = '0.00'
    # 差异
    ws7.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").border = thin_border
    ws7.cell(row=row_idx, column=4).number_format = '0.00'
    ws7.cell(row=row_idx, column=4).fill = calc_fill
    # 差异率
    ws7.cell(row=row_idx, column=5, value=f"=D{row_idx}/B{row_idx}").border = thin_border
    ws7.cell(row=row_idx, column=5).number_format = '0.0%'
    ws7.cell(row=row_idx, column=5).fill = calc_fill
    # 分析
    ws7.cell(row=row_idx, column=6, value=ana).border = thin_border

# 毛利率对比
ws7['A11'] = "毛利率对比"
ws7['A11'].font = Font(name="微软雅黑", size=12, bold=True)

headers9 = ["产品型号", "传统方法", "ABC方法", "差异", "决策影响"]
for i, header in enumerate(headers9, 1):
    cell = ws7.cell(row=12, column=i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

trad_margins = [0.094, -0.134, 0.435, 0.776, 0.861]
abc_margins = [0.167, -0.042, 0.386, 0.590, 0.700]
impact = ["比预期更好", "没那么糟", "基本一致", "被高估了", "被显著高估!"]

for row_idx, (prod, tm, am, imp) in enumerate(zip(products, trad_margins, abc_margins, impact), 13):
    ws7.cell(row=row_idx, column=1, value=prod[1]).border = thin_border
    ws7.cell(row=row_idx, column=2, value=tm).border = thin_border
    ws7.cell(row=row_idx, column=2).number_format = '0.0%'
    ws7.cell(row=row_idx, column=3, value=am).border = thin_border
    ws7.cell(row=row_idx, column=3).number_format = '0.0%'
    ws7.cell(row=row_idx, column=4, value=f"=C{row_idx}-B{row_idx}").border = thin_border
    ws7.cell(row=row_idx, column=4).number_format = '0.0%'
    ws7.cell(row=row_idx, column=4).fill = calc_fill
    ws7.cell(row=row_idx, column=5, value=imp).border = thin_border

# 关键发现
ws7['A19'] = "关键发现："
ws7['A19'].font = Font(name="微软雅黑", size=12, bold=True, color="C00000")

findings = [
    "1. 传统方法严重低估小批量定制品成本!",
    "2. P005真实成本是传统方法的2.16倍!",
    "3. 大批量标准品P001/P002成本被高估约8%",
    "4. P001才是真正的利润贡献主力",
    "5. P005虽高端,但成本极高,要控制规模"
]

for i, finding in enumerate(findings):
    ws7[f'A{20+i}'] = finding
    ws7[f'A{20+i}'].font = Font(name="微软雅黑", size=10, color="C00000")

# 设置列宽
for col in range(1, 7):
    set_column_width(ws7, col, 18)

# ============================================================
# 工作表8: 可视化图表
# ============================================================
ws8 = wb.create_sheet("可视化图表")

ws8['A1'] = "ABC成本模型可视化分析"
ws8['A1'].font = title_font

# 创建成本对比图表数据
ws8['A3'] = "产品单位成本对比（元）"
ws8['A3'].font = Font(name="微软雅黑", size=11, bold=True)

# 数据表
chart_headers = ["产品型号", "传统方法", "ABC方法"]
for i, header in enumerate(chart_headers, 1):
    ws8.cell(row=4, column=i, value=header).font = header_font

for row_idx, (prod, trad, abc) in enumerate(zip(products, traditional_costs, abc_costs), 5):
    ws8.cell(row=row_idx, column=1, value=prod[1])
    ws8.cell(row=row_idx, column=2, value=trad)
    ws8.cell(row=row_idx, column=3, value=abc)

# 创建簇状柱状图
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "传统方法 vs ABC方法 单位成本对比"
chart1.y_axis.title = '单位成本（元）'
chart1.x_axis.title = '产品型号'

data = Reference(ws8, min_col=2, min_row=4, max_row=9, max_col=3)
cats = Reference(ws8, min_col=1, min_row=5, max_row=9)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws8.add_chart(chart1, "E3")

# 作业成本分布数据
ws8['A20'] = "作业成本分布"
ws8['A20'].font = Font(name="微软雅黑", size=11, bold=True)

pie_headers = ["作业类别", "成本金额(元)"]
for i, header in enumerate(pie_headers, 1):
    ws8.cell(row=21, column=i, value=header).font = header_font

activity_categories = [
    ["单位级作业", 4560000],
    ["批次级作业", 2040000],
    ["产品级作业", 620000],
    ["设施级作业", 1080000],
]

for row_idx, cat in enumerate(activity_categories, 22):
    ws8.cell(row=row_idx, column=1, value=cat[0])
    ws8.cell(row=row_idx, column=2, value=cat[1])

# 创建饼图
pie = PieChart()
pie.title = "作业成本分布"
labels = Reference(ws8, min_col=1, min_row=22, max_row=25)
data = Reference(ws8, min_col=2, min_row=21, max_row=25)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.height = 10
pie.width = 15
ws8.add_chart(pie, "E20")

print("工作表 '可视化图表' 创建完成...")

# ============================================================
# 保存文件
# ============================================================
output_file = "瓦轴集团ABC成本模型_演示版.xlsx"
wb.save(output_file)
print(f"\n✓ Excel模型创建成功!")
print(f"✓ 文件保存为: {output_file}")
print(f"\n模型包含以下工作表:")
print("  1. 说明 - 使用说明和项目信息")
print("  2. 基础数据 - 产品信息和生产数据")
print("  3. 成本归集 - 制造费用明细")
print("  4. 作业识别 - 20个作业清单")
print("  5. 成本动因 - 动因选择和分配率")
print("  6. 产品成本(ABC) - ABC方法完全成本")
print("  7. 成本对比 - 传统vs ABC对比分析")
print("  8. 可视化图表 - 成本对比图表")
print(f"\n核心发现:")
print("  • P005真实成本1,649.33元，传统方法仅763.17元，低估116%!")
print("  • P001毛利率16.7%，比传统方法显示的9.4%更好")
print("  • 小批量定制品成本被严重低估，影响定价和决策")
print(f"\n请使用Excel打开文件查看完整模型。")

